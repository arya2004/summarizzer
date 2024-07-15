import textract
import re
from openpyxl import Workbook

def extract_text_from_doc(file_path):
    try:
        text = textract.process(file_path)
        return text.decode('utf-8')
    except Exception as e:
        raise RuntimeError(f"An error occurred while processing the DOC file: {e}")

def parse_subpoints(text):
    subpoints = []
    main_point_pattern = re.compile(r'^\d+\.\d+')
    subpoint_pattern = re.compile(r'^[a-z]\)')
    nested_subpoint_pattern = re.compile(r'^\s*[ivxlc]+\)')
    table_pattern = re.compile(r'((?:\w+\t)+\w+\n)+')  # Adjust this pattern as necessary
    
    lines = text.split('\n')
    current_subpoint = None
    
    for line in lines:
        line = line.strip()
        if main_point_pattern.match(line):
            current_main_point = line
        elif subpoint_pattern.match(line):
            current_subpoint = line
            subpoints.append((current_subpoint, []))
        elif nested_subpoint_pattern.match(line):
            if current_subpoint:
                subpoints[-1][1].append(line)
        else:
            # Check for tables within subpoints
            if current_subpoint and table_pattern.match(line):
                subpoints[-1][1].append(line)

    return subpoints

def store_in_excel(subpoints, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Subpoints"

    ws.cell(row=1, column=1, value="Subpoint")
    row = 2
    for subpoint, nested_subpoints in subpoints:
        full_text = subpoint
        if nested_subpoints:
            full_text += "\n" + "\n".join(nested_subpoints)
        ws.cell(row=row, column=1, value=full_text)
        row += 1
        print(row, full_text)

    wb.save(output_file)

# Path to your DOC file
file_path = 'sotr.doc'
output_file = 'extracted_subpoints_v13.xlsx'

# Extract text
try:
    extracted_text = extract_text_from_doc(file_path)
    subpoints = parse_subpoints(extracted_text)
    store_in_excel(subpoints, output_file)
    print(f"Subpoints have been extracted and stored in {output_file}")
except Exception as e:
    print(e)
